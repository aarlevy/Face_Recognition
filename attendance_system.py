import os
import cv2
import numpy as np
import face_recognition
from datetime import datetime, time
import pandas as pd
import logging
import time as tm
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class AttendanceSystem:
    def __init__(self, database_path='D:\\FaceRecog-Cursor\\Face_Recognition\\Ref-Faces', 
                 attendance_log='attendance_records.xlsx'):
        """
        Initialize the attendance system
        database_path: Path to the directory containing reference face images
        attendance_log: Path to the Excel file storing attendance records
        """
        self.database_path = database_path
        self.attendance_log = attendance_log
        self.known_face_encodings = []
        self.known_face_names = []
        self.face_recognition_tolerance = 0.5
        self.min_confidence = 0.4
        self.cutoff_time = None
        self.today_attendance = {}  # Track today's attendance
        
        # Get local timezone
        self.timezone = pytz.timezone('Africa/Johannesburg')  # Using South African Standard Time (SAST)
        
        # Face detection cascade classifier
        cascade_path = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
        self.face_cascade = cv2.CascadeClassifier(cascade_path)
        
        # Recognition settings
        self.recognition_cooldown = 1.0
        self.last_recognition_time = 0
        self.name_display_duration = 1.5
        self.active_displays = {}

    def set_cutoff_time(self):
        """Set the cutoff time for late attendance"""
        print(f"\nCurrent timezone: {self.timezone}")
        current_time = datetime.now(self.timezone).strftime('%H:%M')
        print(f"Current time: {current_time}")
        
        while True:
            try:
                cutoff_input = input("\nEnter cutoff time (HH:MM in 24-hour format, e.g. 09:00): ")
                hours, minutes = map(int, cutoff_input.split(':'))
                if 0 <= hours <= 23 and 0 <= minutes <= 59:
                    self.cutoff_time = time(hours, minutes)
                    print(f"Cutoff time set to {self.cutoff_time.strftime('%H:%M')}")
                    return
                else:
                    print("Invalid time format. Please use HH:MM in 24-hour format.")
            except ValueError:
                print("Invalid input. Please use HH:MM format (e.g. 09:00)")

    def load_known_faces(self):
        """Load and encode all known faces from the database directory"""
        logger.info("Loading known faces...")
        
        for person_name in os.listdir(self.database_path):
            person_dir = os.path.join(self.database_path, person_name)
            if not os.path.isdir(person_dir):
                continue
                
            logger.info(f"Processing images for {person_name}")
            person_encodings = []
            
            for image_file in os.listdir(person_dir):
                if not any(image_file.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png']):
                    continue
                    
                image_path = os.path.join(person_dir, image_file)
                try:
                    # Load image and find face encodings
                    image = face_recognition.load_image_file(image_path)
                    # Convert to RGB if needed
                    if len(image.shape) == 2:
                        image = cv2.cvtColor(image, cv2.COLOR_GRAY2RGB)
                    elif image.shape[2] == 4:
                        image = cv2.cvtColor(image, cv2.COLOR_BGRA2RGB)
                    
                    # Get face locations first
                    face_locations = face_recognition.face_locations(image, model="hog")
                    if face_locations:
                        # Get encodings for the first face found
                        face_encoding = face_recognition.face_encodings(image, face_locations)[0]
                        person_encodings.append(face_encoding)
                        logger.info(f"Successfully processed {image_file}")
                    else:
                        logger.warning(f"No face found in {image_file}")
                except Exception as e:
                    logger.error(f"Error processing {image_file}: {str(e)}")
            
            # Add all encodings for this person
            if person_encodings:
                self.known_face_encodings.extend(person_encodings)
                self.known_face_names.extend([person_name] * len(person_encodings))
        
        logger.info(f"Loaded {len(self.known_face_encodings)} face encodings for {len(set(self.known_face_names))} people")

    def detect_faces(self, frame):
        """Quick face detection using Haar Cascade"""
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = self.face_cascade.detectMultiScale(
            gray,
            scaleFactor=1.1,
            minNeighbors=5,
            minSize=(30, 30)
        )
        return faces

    def recognize_face(self, frame, face_location):
        """Recognize a face in the given location"""
        x, y, w, h = face_location
        # Extract face region and convert to RGB
        face_region = frame[y:y+h, x:x+w]
        rgb_face = cv2.cvtColor(face_region, cv2.COLOR_BGR2RGB)
        
        # Get face encoding
        face_locations = [(0, w, h, 0)]  # Convert to face_recognition format
        face_encodings = face_recognition.face_encodings(rgb_face, face_locations)
        
        if not face_encodings:
            return None, None
            
        face_encoding = face_encodings[0]
        
        # Compare with known faces
        matches = face_recognition.compare_faces(
            self.known_face_encodings, 
            face_encoding, 
            tolerance=self.face_recognition_tolerance
        )
        face_distances = face_recognition.face_distance(self.known_face_encodings, face_encoding)
        
        if True in matches and min(face_distances) < (1 - self.min_confidence):
            best_match_idx = np.argmin(face_distances)
            if matches[best_match_idx]:
                name = self.known_face_names[best_match_idx]
                confidence = (1 - face_distances[best_match_idx]) * 100
                return name, confidence
        
        return None, None

    def create_excel_report(self):
        """Create a formatted Excel report of today's attendance"""
        wb = Workbook()
        ws = wb.active
        current_date = datetime.now(self.timezone).strftime('%Y-%m-%d')
        ws.title = f"Attendance {current_date}"

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        headers = ["Name", "Arrival Time", "Status", "Confidence"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Write data
        row = 2
        sorted_attendance = sorted(self.today_attendance.items(), key=lambda x: x[1]['time'])
        
        for name, info in sorted_attendance:
            ws.cell(row=row, column=1, value=name).border = border
            ws.cell(row=row, column=2, value=info['time']).border = border
            ws.cell(row=row, column=3, value=info['status']).border = border
            ws.cell(row=row, column=4, value=f"{info['confidence']:.1f}%").border = border
            
            # Color coding for status
            status_cell = ws.cell(row=row, column=3)
            if info['status'] == "Late":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            row += 1

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Add summary
        ws.cell(row=row + 1, column=1, value="Summary").font = Font(bold=True)
        ws.cell(row=row + 2, column=1, value=f"Total Attendance: {len(self.today_attendance)}")
        ws.cell(row=row + 3, column=1, value=f"On Time: {sum(1 for info in self.today_attendance.values() if info['status'] == 'On Time')}")
        ws.cell(row=row + 4, column=1, value=f"Late: {sum(1 for info in self.today_attendance.values() if info['status'] == 'Late')}")
        ws.cell(row=row + 5, column=1, value=f"Cutoff Time: {self.cutoff_time.strftime('%H:%M')}")

        # Save the file
        report_path = f"attendance_report_{current_date}.xlsx"
        wb.save(report_path)
        logger.info(f"Attendance report saved to {report_path}")
        return report_path

    def mark_attendance(self, name, confidence):
        """Record attendance"""
        current_time = datetime.now(self.timezone)
        current_time_str = current_time.strftime('%H:%M:%S')
        
        # Determine if attendance is late or on time
        attendance_status = "Late" if current_time.time() > self.cutoff_time else "On Time"
        
        # Only record first appearance of the day
        if name not in self.today_attendance:
            self.today_attendance[name] = {
                'time': current_time_str,
                'status': attendance_status,
                'confidence': confidence
            }
            return True
        return False

    def run_recognition(self):
        """Run real-time face recognition for attendance"""
        if not self.known_face_encodings:
            logger.error("No faces loaded! Please load faces first.")
            return
            
        if self.cutoff_time is None:
            self.set_cutoff_time()
            
        logger.info("Starting face recognition...")
        logger.info("Press 'q' or 'Q' to quit, or click the window's X button to close and generate the attendance report.")
        
        # Initialize camera
        video_capture = cv2.VideoCapture(0)
        if not video_capture.isOpened():
            logger.error("Could not open webcam")
            return
        
        video_capture.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        video_capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        video_capture.set(cv2.CAP_PROP_FPS, 30)
        
        last_detection = {}
        attendance_cooldown = 5
        
        window_name = 'Attendance System (Press Q to quit)'
        cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
        
        try:
            while True:
                ret, frame = video_capture.read()
                if not ret:
                    logger.error("Failed to grab frame")
                    break
                
                faces = self.detect_faces(frame)
                current_time = tm.time()
                
                # Update active displays and remove expired ones
                expired_displays = []
                for face_key, display_info in self.active_displays.items():
                    if current_time - display_info['start_time'] > self.name_display_duration:
                        expired_displays.append(face_key)
                for face_key in expired_displays:
                    self.active_displays.pop(face_key)
                
                for (x, y, w, h) in faces:
                    face_key = f"{x},{y}"
                    
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 255, 0), 2)
                    
                    should_recognize = (
                        current_time - self.last_recognition_time >= self.recognition_cooldown and
                        face_key not in self.active_displays
                    )
                    
                    if should_recognize:
                        name, confidence = self.recognize_face(frame, (x, y, w, h))
                        
                        if name and confidence:
                            confidence_str = f"{confidence:.1f}%"
                            
                            # Only log if confidence is above 55%
                            if confidence > 55:
                                print(f"✓ Recognized: {name} (Confidence: {confidence_str})")
                            
                            if (name not in last_detection or 
                                current_time - last_detection[name] > attendance_cooldown):
                                if self.mark_attendance(name, confidence):
                                    logger.info(f"Attendance marked for {name} with confidence {confidence_str}")
                                last_detection[name] = current_time
                            
                            self.active_displays[face_key] = {
                                'name': name,
                                'confidence': confidence,
                                'start_time': current_time
                            }
                        else:
                            print("× No match found for detected face")
                        
                        self.last_recognition_time = current_time
                    
                    if face_key in self.active_displays:
                        display_info = self.active_displays[face_key]
                        name = display_info['name']
                        confidence = display_info['confidence']
                        confidence_str = f"{confidence:.1f}%"
                        
                        time_displayed = current_time - display_info['start_time']
                        alpha = min(1.0, (self.name_display_duration - time_displayed) / 1.0)
                        
                        color = (0, int(255 * alpha), 0) if confidence >= 70 else (0, int(255 * alpha), int(255 * alpha))
                        cv2.rectangle(frame, (x, y-35), (x+w, y), color, cv2.FILLED)
                        cv2.putText(frame, f"{name} ({confidence_str})", 
                                  (x + 6, y - 6),
                                  cv2.FONT_HERSHEY_DUPLEX, 0.6, (255, 255, 255), 1)
                    else:
                        cv2.putText(frame, "Unknown", 
                                  (x + 6, y - 6),
                                  cv2.FONT_HERSHEY_DUPLEX, 0.6, (255, 255, 255), 1)
                
                # Add quit instructions to the frame
                cv2.putText(frame, "Press 'Q' to quit", 
                          (10, 30),
                          cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
                
                cv2.imshow(window_name, frame)
                
                # Check for both lowercase and uppercase 'q'
                key = cv2.waitKey(1) & 0xFF
                if key == ord('q') or key == ord('Q') or cv2.getWindowProperty(window_name, cv2.WND_PROP_VISIBLE) < 1:
                    print("\nClosing the program...")
                    break
                    
        finally:
            video_capture.release()
            cv2.destroyAllWindows()
            # Force close all windows
            for i in range(5):
                cv2.waitKey(1)
                cv2.destroyAllWindows()
            
            # Generate Excel report after closing
            if self.today_attendance:
                report_path = self.create_excel_report()
                print(f"\nAttendance report has been saved to: {report_path}")
            else:
                print("\nNo attendance records to report.")

def main():
    system = AttendanceSystem()
    system.load_known_faces()
    system.run_recognition()

if __name__ == "__main__":
    main() 